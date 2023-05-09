from openpyxl import load_workbook

fn = 'main.xlsx'
wb = load_workbook(fn)
ws = wb['Sheet1']
ws['A5'] = 'hello world'

ws.append(['1','2','3'])
ws.append(['4','5','6'])
ws.append(['7','8','9'])
wb.save(fn)
wb.close()
